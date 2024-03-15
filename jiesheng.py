import cv2
from skimage.metrics import structural_similarity as ssim
import os
import openpyxl as op

# 计算两个图像的结构相似性指数
def pic_ssim(img1_path, img2_path):
    # 读取两个图像
    img1 = cv2.imread(img1_path)
    img2 = cv2.imread(img2_path)

    if img1 is None:
        raise FileNotFoundError(f"Image file not found: {img1_path}")

    if img2 is None:
        raise FileNotFoundError(f"Image file not found: {img2_path}")

    # 将图像转换为灰度
    img1_gray = cv2.cvtColor(img1, cv2.COLOR_BGR2GRAY)
    img2_gray = cv2.cvtColor(img2, cv2.COLOR_BGR2GRAY)

    # 计算两个图像的结构相似性指数
    similarity_index = ssim(img1_gray, img2_gray)

    # 如果相似度大于或等于0.8，返回True
    if similarity_index >= 0.8:
        return similarity_index, True

    # 否则，返回False
    return similarity_index, False


# 写一个方法读取txt文件中的json数据，并提取所有json数据中的usage里的总步数/总探索token数/总反射token数
def read_json_from_txt(txt_path1, txt_path2):
    step_num = 0
    explore_tokens = 0
    reflect_tokens = 0
    with open(txt_path1, 'r', encoding='utf-8') as f:
        lines = f.readlines()
        step_num = len(lines)
        for line in lines:
            line = line.strip()
            if line:
                line = eval(line)
                explore_tokens += line['response']['usage']['total_tokens']
    with open(txt_path2, 'r', encoding='utf-8') as f:
        lines = f.readlines()
        for line in lines:
            line = line.strip()
            if line:
                line = eval(line)
                reflect_tokens += line['response']['usage']['total_tokens']
    return step_num, explore_tokens, reflect_tokens


# 写一个函数，传入参数为num,相关数据和excel的地址,将step_num, explore_tokens, reflect_tokens, similarity_index写入excel表格的num行
def write_to_excel(num, step_num, explore_tokens, reflect_tokens, similarity_index, excel_path):
    tableAll = op.load_workbook(excel_path)
    table1 = tableAll['Sheet1']
    table1.cell(num, 3, explore_tokens)
    table1.cell(num, 4, reflect_tokens)
    table1.cell(num, 5, similarity_index)
    table1.cell(num, 6, step_num)
    tableAll.save(excel_path)




def learn(app, task, using_method, num):
    # 执行学习脚本
    os.system(f"python learn.py --app {app} --task {task} --using_method {using_method} --num {num}")


# 写一个函数，传入参数为num和excel的地址，功能是从excel表格中读取第num行的第一列的app和第二列task的值
def learn_from_excel(num, excel_path):
    workbook = op.load_workbook(excel_path)
    # 选择要读取的工作表
    worksheet = workbook['Sheet1']
    row = worksheet[num]
    app = row[0].value
    task = '"' + row[1].value + '"'
    using_method = row[2].value
    workbook.close()
    return app, task, using_method


if __name__ == '__main__':
    num = 5
    print("jiesheng:excel表格中的第" + str(
        num) + "行任务开始执行了!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
    # 从excel中读取app和task
    app, task, using_method = learn_from_excel(num, "./task_result.xlsx")
    # 调用learn执行第一次任务,self_explorer.py中将相关数据存入excel表格
    learn(app, task, using_method, num)
